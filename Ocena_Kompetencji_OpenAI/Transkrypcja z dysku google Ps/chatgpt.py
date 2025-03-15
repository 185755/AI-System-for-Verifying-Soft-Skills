import openai
import os
import pandas as pd
from openpyxl import load_workbook
from glob import glob

a = 1  # Pierwszy indeks
b = 26  # Ostatni indeks

openai.api_key = os.getenv('OPENAI_API_KEY')

# Wczytanie danych z Excela
df = pd.read_excel('profil_kompetencji_dla_weryfikacji_filmów.xlsx', header=None).transpose()
opis_kompetencji = df.iloc[0].dropna().values
pytania_rekrutacyjne = df.iloc[1].dropna().values
wskazniki_oceny = df.iloc[2].dropna().values

# Folder z transkrypcjami
transcription_folder = "./transkrypcje/zad3"  # Ścieżka do folderu z transkrypcjami
transcription_files = glob(os.path.join(transcription_folder, "*.txt"))

# Ścieżka do folderu wynikowego
output_directory = os.path.join("oceny kompetencji", "Zad3-o1-mini")
os.makedirs(output_directory, exist_ok=True)

for transcription_file in transcription_files:

    file_name = os.path.splitext(os.path.basename(transcription_file))[0]
    with open(transcription_file, 'r', encoding='utf-8') as f:
        transcription_text = f.read()

    os.makedirs(output_directory, exist_ok=True)
    output_file = os.path.join(output_directory, f"{file_name}.xlsx") ###Tutaj zmień!!!!!!!!!!!!!!!!!!!!

    # Sprawdzenie czy plik już istnieje
    file_exists = os.path.exists(output_file)

    # Jeżeli plik nie istnieje, tworzymy nowy DataFrame z odpowiednimi kolumnami
    if not file_exists:
        df_result = pd.DataFrame(columns=["Indeks", "Opis Kompetencji", "Pytanie Rekrutacyjne", "Wskaźnik Oceny", "Odpowiedź", "Ocena", "Tokeny użyte", "Ocena Asesora"])
    else:
        # Jeżeli plik istnieje, wczytujemy go
        df_result = pd.read_excel(output_file)

    # Iteracja przez zadane indeksy
    for i in range(a, b):
        question = (
            f'Transkrypcja odpowiadała na pytanie rekrutacyjne, które brzmiało: "{pytania_rekrutacyjne[i]}". '
            f'Na podstawie transkrypcji zdecyduj, czy osoba wykazała się {wskazniki_oceny[i]}, '
            f'jeżeli tak, to podaj fragmenty potwierdzające to i przedstaw argumenty. '
            f'Na koniec oceń {wskazniki_oceny[i]} osoby (1 jeżeli osoba wykazała się daną cechą, lub 0 jeżeli się nie wykazała). '
            f'Nie bądź zbyt łagodny w swojej ocenie. Ostatnim znakiem twojej odpowiedzi nie może być nic innego niż ostateczna ocena (0 lub 1).'
        )

        messages = [
            {"role": "user", "content": f"Tutaj masz transkrypcje tekstu:\n\n{transcription_text}\n\n{question}"}
        ]

        response = openai.chat.completions.create(
            messages=messages,
            model="o1-mini",
            seed=123,
            temperature=1,
        )

        # Odczytaj odpowiedź i liczbę tokenów
        response_text = response.choices[0].message.content
        tokens = response.usage.total_tokens  

        print(response_text)

        ocena = response_text[-1] if response_text[-1] in ['0', '1'] else response_text[-2]

        # Dodanie nowego wiersza do DataFrame
        new_row = {
            "Indeks": i,
            "Opis Kompetencji": opis_kompetencji[i],
            "Pytanie Rekrutacyjne": pytania_rekrutacyjne[i],
            "Wskaźnik Oceny": wskazniki_oceny[i],
            "Odpowiedź": response_text,
            "Ocena": ocena,
            "Tokeny użyte": tokens
        }
        df_result = df_result._append(new_row, ignore_index=True)

        print(f"Odpowiedź została zapisana w DataFrame dla indeksu {i}.")

    # Zapis do pliku Excel
    df_result.to_excel(output_file, index=False)
    print(f"Plik Excel został zapisany jako '{output_file}'.")


    # Teraz otwieramy plik Excel, aby dodać metryki i formuły
    wb = load_workbook(output_file)
    ws = wb.active  # Zakładam, że to pierwszy arkusz

    last_row = ws.max_row
    ws['I1'] = "Metryki"
    ws['I2'] = "TP (True Positive)"
    ws['I3'] = "FP (False Positive)"
    ws['I4'] = "FN (False Negative)"
    ws['I5'] = "TN (True Negative)"
    ws['I6'] = "Precision"
    ws['I7'] = "Recall"
    ws['I8'] = "F1-Score"
    ws['I9'] = "Accuracy"

    ws['J2'] = f'=COUNTIFS(F2:F{last_row}, 1, H2:H{last_row}, 1)'  # TP
    ws['J3'] = f'=COUNTIFS(F2:F{last_row}, 1, H2:H{last_row}, 0)'  # FP
    ws['J4'] = f'=COUNTIFS(F2:F{last_row}, 0, H2:H{last_row}, 1)'  # FN
    ws['J5'] = f'=COUNTIFS(F2:F{last_row}, 0, H2:H{last_row}, 0)'  # TN
    ws['J6'] = f'=IF(J2+J3=0, 0, J2/(J2+J3))'  # Precision
    ws['J7'] = f'=IF(J2+J4=0, 0, J2/(J2+J4))'  # Recall
    ws['J8'] = f'=IF(J5+J6=0, 0, 2*J7*J6/(J7+J6))'  # F1-Score
    ws['J9'] = f'=IF(J2+J5=0, 0, (J2+J5)/(J2+J4+J3+J5))'

    # Zapisujemy zmiany
    wb.save(output_file)
    print("Dodano metryki oraz formuły do pliku Excel.")